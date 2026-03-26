import time
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# --- GUI 라이브러리 ---
import tkinter as tk
from tkinter import ttk  # '세련된' 위젯을 위해
from tkinter import font as tkFont # 폰트
from tkinter import messagebox # 팝업 메시지
from tkinter import filedialog # '파일 열기' 및 '다른 이름으로 저장'

# --- 스레딩 라이브러리 (GUI 멈춤 방지) ---
import threading
import sys # .exe 경로 탐지를 위해

# ======================================================================
# (1) 사용자 설정 (기존 코드)
# ======================================================================

# (삭제) --- EXCEL_FILE_NAME 변수 삭제 ---
# 이제 파일 경로를 사용자가 직접 입력받습니다.

# 2. 검색어(B열)가 시작되는 행 번호 (1, 2행 제외)
START_ROW = 3

# 3. API 기본 URL (상수)
BASE_URL = "https://biologicalvariation.eu/api/search/by_analyte"

# ======================================================================
# (2) 헬퍼 함수 (기존 코드)
# ======================================================================

def to_float(value_text):
    """텍스트를 float(소수점 숫자)로 변환, 실패 시 원본 텍스트 반환"""
    try:
        return float(value_text)
    except ValueError:
        return value_text # 숫자로 바꿀 수 없으면 원본 텍스트 반환

# ======================================================================
# (3) 단일 시트 처리 함수 (기존 코드 - 수정 없음)
# ======================================================================

def process_sheet(sheet, driver, wait):
    """
    하나의 엑셀 시트(sheet)를 받아 스크래핑 작업을 수행합니다.
    (수정됨: 시트별 종료 행 지정, 오류 메시지 변경, 빈 셀 메시지 제거)
    """
    
    # --- B1 셀에 스크립트 실행 시간 기록 ---
    try:
        update_time_str = time.strftime("%Y-%m-%d %H:%M:%S")
        sheet.cell(row=1, column=2).value = update_time_str # B1셀
        print(f"[{sheet.title}] 시트 B1 셀에 실행 시간 기록: {update_time_str}")
    except Exception as e:
        print(f"!!! 오류: [{sheet.title}] 시트 B1 셀에 시간 기록 실패: {e}")

    # (신규) --- 시트별 종료 행 설정 ---
    sheet_title = sheet.title
    end_row = 0 
    
    if sheet_title == 'Chemistry':
        end_row = 175
        print(f"  [{sheet_title}] 시트의 종료 행을 175행으로 설정합니다.")
    elif sheet_title == 'Immunology':
        end_row = 111
        print(f"  [{sheet_title}] 시트의 종료 행을 111행으로 설정합니다.")
    else:
        # 지정되지 않은 시트의 경우, 기존 로직(max_row) 사용
        print(f"  [{sheet_title}] 시트의 종료 행이 지정되지 않아, 시트의 최대 행({sheet.max_row})까지 실행합니다.")
        end_row = sheet.max_row 
    
    print(f"  [{sheet_title}] 시트 작업을 시작합니다. (시작: {START_ROW}행, 종료: {end_row}행)")

    # --- B열을 순회하며 검색어 추출 및 스크래핑 ---
    # (수정) `end_row + 1`을 사용하여 설정된 종료 행을 *포함*하도록 함
    for row in range(START_ROW, end_row + 1):
        search_term_cell = sheet.cell(row=row, column=2) # B열 = column 2
        search_term = search_term_cell.value

        # (수정) --- B열이 비었을 때 C열에 기록하는 코드 삭제 ---
        if not search_term or (isinstance(search_term, str) and search_term.strip() == ""):
            print(f"  [{sheet.title}] B{row} 셀에 검색어가 없습니다. 다음 행으로 넘어갑니다.")
            # sheet.cell(row=row, column=3).value = "EFLM 검색어 없음" # <-- 요청에 따라 삭제
            continue # 다음 행(row)으로 바로 이동
            
        # B열 값이 "검색어 확인 불가"이면 다음 행으로 건너뛰기
        if isinstance(search_term, str) and search_term.strip() == "검색어 확인 불가":
            print(f"  [{sheet.title}] B{row} 셀이 '검색어 확인 불가'입니다. 다음 행으로 넘어갑니다.")
            continue # 다음 행(row)으로 바로 이동
        
        print(f"\n--- [ {sheet.title} / {row}행 ] '{search_term}' 처리 중 ---")

        try:
            # 1. 검색 URL로 이동
            full_url = f"{BASE_URL}?format=html&query={search_term}"
            driver.get(full_url)
            
            # "No Meta-Analysis Results" 확인 로직
            wait.until(EC.visibility_of_element_located((By.TAG_NAME, "body")))
            no_results = driver.find_elements(By.XPATH, "//*[contains(text(), 'No Meta-Analysis Results')]")

            if len(no_results) > 0: # "No Meta-Analysis Results" 텍스트를 찾았다면
                print(f"  'No Meta-Analysis Results' 발견. C{row}에 기록합니다.")
                sheet.cell(row=row, column=3).value = "No Meta-Analysis Results"
                continue # --- 다음 행(row)으로 즉시 이동 ---

            # --- "No Meta-Analysis Results"가 없을 때만 아래 코드 실행 ---

            # (1) 검색 결과 목록 페이지 처리 (2단계 매칭)
            try:
                # 0.5초의 짧은 대기시간으로 'Analytical...' 버튼을 찾아봄
                WebDriverWait(driver, 0.5).until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(text(), 'Analytical Performance Specification')]")
                ))
                # 버튼을 찾았다면(성공) -> 상세 페이지이므로 아무것도 안함
                print("  (정보) 상세 페이지로 바로 이동했습니다.")
            except:
                # 버튼을 못 찾았다면(Timeout) -> 검색 목록 페이지로 간주
                print(f"  검색 목록 페이지로 판단. '{search_term}'(와)과 일치하는 항목을 찾습니다.")
                
                try:
                    # 1. B열의 검색어(search_term)와 텍스트가 *정확히* 일치하는 버튼을 찾습니다.
                    print(f"  1차 시도: '{search_term}'(와)과 정확히 일치하는 버튼 검색...")
                    exact_match_button = wait.until(EC.element_to_be_clickable(
                        (By.XPATH, f"//button[normalize-space()='{search_term}']")
                    ))
                    
                    print(f"  '{search_term}' 버튼 클릭.")
                    exact_match_button.click()

                except Exception as e:
                    # 2. 정확한 일치에 실패하면, 괄호() 앞부분으로 다시 시도합니다.
                    print(f"  1차 시도 실패.")
                    main_term = "" # main_term 변수 초기화
                    try:
                        main_term = search_term.split('(')[0].strip()
                        
                        # 괄호가 없었거나(main_term == search_term), 괄호를 잘라도 빈 문자열이면
                        if main_term == search_term or not main_term:
                            print(f"  부분 일치로 찾을 검색어가 없습니다. 건너뜁니다.")
                            raise e # 1차 시도의 Exception을 다시 발생시킴 (e2로 잡힘)

                        print(f"  2차 시도: 괄호를 제외한 '{main_term}'(와)과 정확히 일치하는 버튼 검색...")
                        
                        partial_match_button = wait.until(EC.element_to_be_clickable(
                            (By.XPATH, f"//button[normalize-space()='{main_term}']")
                        ))
                        
                        print(f"  '{main_term}' 버튼 클릭.")
                        partial_match_button.click()

                    except Exception as e2:
                        # 1차, 2차 시도 모두 실패
                        print(f"  !!! 오류: 검색 목록에서 '{search_term}' 또는 '{main_term}'(을)를 찾지 못했습니다.")
                        sheet.cell(row=row, column=3).value = "Search List Error"
                        continue # 다음 행(row)으로 이동

                # --- (공통) 버튼 클릭 후 상세 페이지 로딩 대기 ---
                try:
                    # 상세 페이지로 넘어간 후, 'Analytical...' 버튼이 로드될 때까지 대기
                    wait.until(EC.visibility_of_element_located(
                        (By.XPATH, "//button[contains(text(), 'Analytical Performance Specification')]")
                    ))
                    print("  상세 페이지 로딩 완료.")
                # (수정) --- 오류 메시지 변경 ---
                except Exception as page_load_e:
                    print(f"  !!! 오류: 버튼은 클릭했으나 상세 페이지 로딩에 실패했습니다.")
                    # "Page Load Error" -> "No Meta-Analysis Results"로 변경
                    sheet.cell(row=row, column=3).value = "No Meta-Analysis Results" 
                    continue # 다음 행(row)으로 이동
            # --- 처리 끝 ---


            # (2) 2. 'Analytical Performance Specification' 버튼 클릭 (기존 로직)
            spec_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(text(), 'Analytical Performance Specification')]")
            ))
            spec_button.click()
            print("  버튼 1: 'Analytical Performance Specification' 클릭")

            # 3. 팝업(모달) 창이 뜰 때까지 대기
            wait.until(EC.visibility_of_element_located(
                (By.XPATH, "//div[@class='modal-content']")
            ))
            print("  팝업 창 확인")

            # 4. 'Calculate' 클릭 전 정보 엑셀 O, P열에 입력
            try:
                # CVI 값 추출 및 입력
                input1_val = driver.find_element(By.XPATH, "//div[@class='modal-content']//label[contains(text(), 'Within-subject')]/following-sibling::input[1]").get_attribute("value")
                sheet.cell(row=row, column=15).value = to_float(input1_val) # O열 (15)
                
                # CVG 값 추출 및 입력
                input2_val = driver.find_element(By.XPATH, "//div[@class='modal-content']//label[contains(text(), 'Between-subject')]/following-sibling::input[1]").get_attribute("value")
                sheet.cell(row=row, column=16).value = to_float(input2_val) # P열 (16)
                
                print(f"  엑셀 {row}행 O, P열에 CVI(={input1_val}), CVG(={input2_val}) 값 입력 완료")
                
            except Exception as e:
                print(f"  !!! 팝업 정보(CVI/CVG)를 가져오거나 엑셀에 입력하는 데 실패했습니다: {e}")


            # 5. 팝업 창의 'Calculate' 버튼 클릭
            calculate_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[@class='modal-content']//button[contains(text(), 'Calculate')]")
            ))
            calculate_button.click()
            print("  버튼 2: 'Calculate' 클릭")
            
            # 6. 'Results' 표가 나타날 때까지 대기
            wait.until(EC.visibility_of_element_located(
                (By.XPATH, "//div[@class='modal-content']//table")
            ))
            print("  결과 표 로드 완료")
            
            # 7. 표 데이터 추출 및 엑셀 C~N열 입력
            try:
                results_table = driver.find_element(By.XPATH, "//div[@class='modal-content']//table")
                
                min_row = results_table.find_element(By.XPATH, ".//tr[contains(td, 'Minimum')]")
                des_row = results_table.find_element(By.XPATH, ".//tr[contains(td, 'Desirable')]")
                opt_row = results_table.find_element(By.XPATH, ".//tr[contains(td, 'Optimal')]")
                
                min_cells = min_row.find_elements(By.TAG_NAME, "td")
                des_cells = des_row.find_elements(By.TAG_NAME, "td")
                opt_cells = opt_row.find_elements(By.TAG_NAME, "td")
                
                # [Optimal] C, D, E, F
                sheet.cell(row=row, column=3).value = to_float(opt_cells[1].text)
                sheet.cell(row=row, column=4).value = to_float(opt_cells[2].text)
                sheet.cell(row=row, column=5).value = to_float(opt_cells[3].text)
                sheet.cell(row=row, column=6).value = to_float(opt_cells[4].text)
                
                # [Desirable] G, H, I, J
                sheet.cell(row=row, column=7).value = to_float(des_cells[1].text)
                sheet.cell(row=row, column=8).value = to_float(des_cells[2].text)
                sheet.cell(row=row, column=9).value = to_float(des_cells[3].text)
                sheet.cell(row=row, column=10).value = to_float(des_cells[4].text)

                # [Minimum] K, L, M, N
                sheet.cell(row=row, column=11).value = to_float(min_cells[1].text)
                sheet.cell(row=row, column=12).value = to_float(min_cells[2].text)
                sheet.cell(row=row, column=13).value = to_float(min_cells[3].text)
                sheet.cell(row=row, column=14).value = to_float(min_cells[4].text)
                
                print(f"  엑셀 {row}행 C열부터 N열까지 12개 데이터 입력 완료")

            except Exception as table_e:
                print(f"  !!! 오류: 표 데이터 추출 또는 엑셀 입력 중 실패")
                print(f"  상세 정보: {table_e}")
                sheet.cell(row=row, column=3).value = "Extraction Error"

        except Exception as e:
            print(f"!!! 오류: '{search_term}' 처리 중 문제가 발생했습니다. 다음 항목으로 넘어갑니다.")
            print(f"상세 정보: {e}")
            sheet.cell(row=row, column=3).value = "Error"
            continue
        
        # *** (중요) 서버 부하 방지를 위한 1초 대기 ***
        time.sleep(1)


# ======================================================================
# (4) 스크래핑 실행 함수 (GUI가 호출할 함수)
# ======================================================================

# (수정) --- 'excel_path'를 인자로 받도록 변경 ---
def start_scraping_logic(sheets_to_process, excel_path):
    """
    기존 main() 함수의 핵심 로직을 수행합니다.
    GUI의 버튼에 의해 호출되며, 별도 스레드에서 실행됩니다.
    """
    
    workbook = None
    driver = None
    
    try:
        # (삭제) --- 1. 엑셀 파일 경로 설정 ---
        # 이 로직은 `run_task` 함수(파일 선택)로 이동했습니다.
            
        # --- 2. 엑셀 파일 열기 (한 번만) ---
        try:
            # (수정) 인자로 받은 'excel_path' 사용
            print(f"엑셀 파일을 엽니다: {excel_path}...")
            workbook = openpyxl.load_workbook(excel_path)
        except FileNotFoundError:
            # (수정) 오류 메시지를 동적으로 변경
            print(f"오류: 엑셀 파일 '{excel_path}'을(를) 찾을 수 없습니다.")
            messagebox.showerror("파일 오류", 
                                 f"오류: 엑셀 파일 '{os.path.basename(excel_path)}'을(를) 찾을 수 없습니다.\n"
                                 f"정확한 파일을 선택했는지 확인하세요.")
            return # 함수 종료
        except Exception as e:
            print(f"오류: 엑셀 파일을 여는 데 실패했습니다. {e}")
            messagebox.showerror("엑셀 오류", f"오류: 엑셀 파일을 여는 데 실패했습니다.\n{e}")
            return # 함수 종료

        # --- 3. Selenium 웹드라이버 설정 (한 번만) ---
        try:
            print("크롬 웹드라이버를 설정합니다 (자동 다운로드)...")
            service = Service(ChromeDriverManager().install())
            options = webdriver.ChromeOptions()
            # options.add_argument('--headless') # 창 숨기기 (주석 해제 시)
            driver = webdriver.Chrome(service=service, options=options)
            wait = WebDriverWait(driver, 10) # 최대 10초 대기
        except Exception as e:
            print(f"오류: 크롬 드라이버 설정에 실패했습니다. {e}")
            messagebox.showerror("드라이버 오류", 
                                 f"오류: 크롬 드라이버 설정에 실패했습니다.\n{e}\n"
                                 "인터넷 연결을 확인하거나, 크롬 브라우저를 업데이트해보세요.")
            return # 함수 종료

        print("스크래핑을 시작합니다...")

        # --- 4. 정의된 모든 시트 순회 ---
        for sheet_name in sheets_to_process:
            print(f"\n=========================================")
            print(f"  시트 처리 시작: [{sheet_name}]")
            print(f"=========================================")
            try:
                # 엑셀에서 해당 시트 가져오기
                sheet = workbook[sheet_name]
                # 시트 처리 함수 호출
                process_sheet(sheet, driver, wait)
                
            except KeyError:
                print(f"!!! 오류: 시트 '{sheet_name}'을(를) 엑셀 파일에서 찾을 수 없습니다.")
                messagebox.showwarning("시트 오류", 
                                       f"오류: 시트 '{sheet_name}'을(를) 엑셀 파일에서 찾을 수 없습니다.\n"
                                       "이 시트를 건너뜁니다.")
            except Exception as e:
                print(f"!!! 치명적인 오류 발생 ({sheet_name} 시트 처리 중): {e}")
                messagebox.showerror("치명적 오류", 
                                     f"!!! 치명적인 오류 발생 ({sheet_name} 시트 처리 중): {e}\n"
                                     "이 시트를 건너뜁니다.")

        # --- 5. 엑셀 파일 '다른 이름으로 저장' 및 브라우저 종료 ---
        
        save_path = "" 
        try:
            # (신규) '다른 이름으로 저장' 시 기본 경로를 설정하기 위해
            # .exe 또는 .py 파일이 있는 폴더 경로를 여기서 계산합니다.
            try:
                if getattr(sys, 'frozen', False):
                    # .exe로 실행된 경우 (frozen)
                    current_folder = os.path.dirname(sys.executable)
                else:
                    # .py 스크립트로 실행된 경우
                    current_folder = os.path.dirname(os.path.abspath(__file__))
            except NameError:
                # 예외 처리 (예: IDLE에서 실행)
                current_folder = os.getcwd()
            
            print("'다른 이름으로 저장' 대화상자를 엽니다...")
            
            default_filename = f"EFLM Autosheet_{time.strftime('%Y-%m-%d')}.xlsx"
            
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialdir=current_folder, 
                initialfile=default_filename, 
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )

            if save_path: # 사용자가 '저장'을 눌렀다면
                workbook.save(save_path)
                print(f"\n--- 작업 완료! ---")
                print(f"모든 데이터를 '{save_path}'에 성공적으로 저장했습니다.")
                
                # 작업 완료 팝업
                messagebox.showinfo("작업 완료", 
                                    f"모든 데이터를 '{save_path}'에 성공적으로 저장했습니다.")
                
                # --- 6. 저장에 성공한 경우에만 파일 열기 ---
                try:
                    print(f"결과 파일을 엽니다...")
                    os.startfile(save_path) 
                except Exception as e:
                    print(f"엑셀 파일을 자동으로 열 수 없습니다. 직접 열어주세요. ({save_path})")
                    messagebox.showwarning("파일 열기 실패", 
                                           f"엑셀 파일을 자동으로 열 수 없습니다.\n직접 열어주세요.\n({save_path})")
            
            else: # 사용자가 '취소'를 눌렀다면
                print(f"\n--- 작업 취소 ---")
                print("파일 저장이 취소되었습니다.")
                messagebox.showwarning("작업 취소", "파일 저장이 취소되었습니다.")
                
        except Exception as e:
            print(f"!!! 오류: 엑셀 파일을 저장하는데 실패했습니다.")
            print(f"상세 정보: {e}")
            messagebox.showerror("저장 오류", f"!!! 오류: 엑셀 파일을 저장하는데 실패했습니다.\n{e}")

    except Exception as general_e:
        # 이 함수 내에서 발생한 모든 예외 처리
        print(f"!!! 스크래핑 로직 실행 중 예외 발생: {general_e}")
        messagebox.showerror("알 수 없는 오류", f"스크래핑 실행 중 오류가 발생했습니다.\n{general_e}")
        
    finally:
        # --- 6. 리소스 정리 ---
        if driver:
            driver.quit()
            print("\n브라우저를 닫았습니다.")
        if workbook:
            workbook.close() # 읽기/쓰기 완료 후 닫기
            print("엑셀 파일을 닫았습니다.")
        
        print("스크래핑 스레드 종료.")


# ======================================================================
# (5) 메인 GUI 실행 함수
# ======================================================================

def main_gui():
    """
    GUI를 생성하고 실행합니다.
    """
    
    # --- 1. 루트 윈도우 생성 ---
    root = tk.Tk()
    root.title("EFLM Data Auto input")

    # --- 2. '세련된' 스타일 및 큼직한 폰트 설정 ---
    style = ttk.Style()
    try:
        style.theme_use('clam') 
    except tk.TclError:
        style.theme_use('default') # 'clam'이 없으면 기본값 사용

    # 큼직한 버튼 폰트
    button_font = tkFont.Font(family='Helvetica', size=16, weight='bold')
    
    # ttk 버튼 스타일 재정의
    style.configure('TButton', font=button_font, padding=(20, 20))

    # --- 3. 윈도우 중앙 정렬 ---
    window_width = 500
    window_height = 400
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    center_x = int(screen_width / 2 - window_width / 2)
    center_y = int(screen_height / 2 - window_height / 2)
    root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
    root.resizable(False, False)

    # --- 4. 메인 프레임 생성 (위젯을 담을 컨테이너) ---
    main_frame = ttk.Frame(root, padding="30 30 30 30")
    main_frame.pack(expand=True, fill=tk.BOTH)

    # --- 5. 버튼 클릭 시 실행될 함수 정의 ---
    
    # 버튼 리스트 (활성화/비활성화를 위해)
    all_buttons = []

    # (수정) --- 'run_task' 함수에 파일 선택 로직 추가 ---
    def run_task(sheets_to_process):
        """
        버튼 클릭 시 호출될 래퍼(wrapper) 함수
        1. (신규) 파일 선택 대화상자 열기
        2. (신규) 파일이 선택되었는지 확인
        3. 모든 버튼 비활성화
        4. 스레드를 생성하여 start_scraping_logic 실행
        5. 스레드 완료 시 버튼을 다시 활성화할 콜백 예약
        """
        
        # 1. 파일 선택 대화상자 열기
        print("사용자에게 엑셀 파일 선택을 요청합니다...")
        
        # (신규) .exe/.py 파일이 있는 폴더를 기본 경로로 설정
        try:
            if getattr(sys, 'frozen', False):
                initial_dir = os.path.dirname(sys.executable)
            else:
                initial_dir = os.path.dirname(os.path.abspath(__file__))
        except NameError:
            initial_dir = os.getcwd()

        # 메인 창을 잠시 숨겨 파일 대화상자에 집중
        root.withdraw()
        
        file_path = filedialog.askopenfilename(
            title="EFLM 엑셀 파일(양식)을 선택하세요",
            initialdir=initial_dir, # .exe 또는 .py가 있는 폴더에서 시작
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        # 파일 대화상자가 닫히면 메인 창 다시 표시
        root.deiconify()

        # 2. 파일이 선택되었는지 확인
        if not file_path:
            print("파일 선택이 취소되었습니다.")
            return # 사용자가 '취소'를 눌렀으므로 아무것도 하지 않음

        print(f"선택된 파일: {file_path}")

        # --- (이하는 기존 로직) ---
        
        # 3. 모든 버튼 비활성화
        for btn in all_buttons:
            btn.config(state='disabled')
        
        # 5. 스레드 완료 시 실행될 콜백 (버튼 활성화)
        def on_complete():
            for btn in all_buttons:
                btn.config(state='normal')
            print("--- 모든 작업 완료. 버튼 활성화 ---")

        # 4. 스레드에서 실행될 실제 작업
        def thread_target():
            # filedialog가 메인 스레드의 Tk 객체를 사용해야 하므로,
            # 스레드 시작 전 Tk 객체를 숨깁니다.
            root.withdraw() # 메인 GUI 숨기기 (filedialog 팝업만 보이도록)
            
            # (수정) 'file_path'를 인자로 전달
            start_scraping_logic(sheets_to_process, file_path) # 시간이 오래 걸리는 작업
            
            # 작업이 끝나면, GUI 관련 작업(버튼 활성화, 창 다시 보이기)을
            # 메인 스레드에서 실행하도록 예약합니다.
            root.after(0, on_complete)
            root.after(100, root.deiconify) # 0.1초 후 메인 GUI 다시 표시

        # 스레드 생성 및 시작
        print(f"\n>>> 스레드 시작: {sheets_to_process} 작업")
        threading.Thread(target=thread_target, daemon=True).start()


    # --- 6. 버튼 생성 및 배치 ---
    
    # 1. ALL 버튼
    btn_all = ttk.Button(
        main_frame, 
        text="ALL (Chemistry, Immunology)", 
        command=lambda: run_task(['Chemistry', 'Immunology'])
    )
    btn_all.pack(expand=True, fill=tk.BOTH, pady=15)

    # 2. Chemistry 버튼
    btn_chem = ttk.Button(
        main_frame, 
        text="Chemistry", 
        command=lambda: run_task(['Chemistry'])
    )
    btn_chem.pack(expand=True, fill=tk.BOTH, pady=15)

    # 3. Immunology 버튼
    btn_immuno = ttk.Button(
        main_frame, 
        text="Immunology", 
        command=lambda: run_task(['Immunology'])
    )
    btn_immuno.pack(expand=True, fill=tk.BOTH, pady=15)

    # 비활성화/활성화를 위해 리스트에 추가
    all_buttons.extend([btn_all, btn_chem, btn_immuno])

    # --- 7. GUI 실행 ---
    root.mainloop()

# ======================================================================
# (6) 스크립트 실행
# ======================================================================

if __name__ == "__main__":
    main_gui()