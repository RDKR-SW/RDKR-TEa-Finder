import openpyxl
import json

file_path = r'C:\Users\심상욱\Desktop\EFLM Autosheet\EFLM Autosheet_2026-02-02.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active # Using the active sheet for now

data = []
# Assuming the columns are 'Test name' and 'Search name'
# Let's find the headers
headers = [cell.value for cell in sheet[1]]
test_idx = -1
search_idx = -1

for i, h in enumerate(headers):
    if h and 'Test name' in str(h):
        test_idx = i + 1
    if h and 'Search name' in str(h):
        search_idx = i + 1

if test_idx == -1 or search_idx == -1:
    # Try row 2 if row 1 is not headers
    headers = [cell.value for cell in sheet[2]]
    for i, h in enumerate(headers):
        if h and 'Test name' in str(h):
            test_idx = i + 1
        if h and 'Search name' in str(h):
            search_idx = i + 1

if test_idx != -1 and search_idx != -1:
    for row in range(3, 15): # Just get a sample of 12 items
        test = sheet.cell(row=row, column=test_idx).value
        search = sheet.cell(row=row, column=search_idx).value
        if test:
            data.append({"test_name": test, "search_name": search})

print(json.dumps(data, ensure_ascii=False))
